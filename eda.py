import matplotlib.pyplot as plt
from sklearn.decomposition import PCA, FastICA
import numpy as np
import seaborn as sns
import matplotlib as mpl
import pandas as pd
import os
from mlxtend.plotting import plot_confusion_matrix
from sklearn.metrics import confusion_matrix
from numpy.linalg import norm
from matplotlib.colors import TwoSlopeNorm
from mpl_toolkits.axes_grid1 import make_axes_locatable
from sklearn.preprocessing import RobustScaler
from sklearn.preprocessing import StandardScaler

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openpyxl.worksheet.datavalidation import DataValidation

from openpyxl.formatting.rule import ColorScaleRule
import matplotlib.pyplot as plt
import numpy as np
import os
import glob
from GeoDS import utilities
import json

def _get_pcs_labels(n_components):
    #For Antoine's update to PC01, PC02, instead use this code f'PC{"{:02d}".format(x+1)}', beware to test all the functionalities if you do so
    pc_lstr = [f'PC{x + 1}' for x in range(n_components)]
    return pc_lstr

def _get_ics_labels(n_components):
    ic_lstr = [f'IC{x + 1}' for x in range(n_components)]
    return ic_lstr


def _autolabel(heights, num_pc, ax):
    """ Private function, Generate cumulative explained variance labels for scree plot.
    """
    i = 1
    previous_height = 0
    for height in heights:
        #PH Note : had to add int(previous_height) < 99 condition so that the last PC showing 100 would not show, in case many PCs would explain about 99 percent of variance.
        if int(height) != int(previous_height) and int(previous_height) < 99:
            ax.annotate(
                int(height),
                xy=(i, height+1.25),
                ha='right',
                va='bottom',
                fontsize=9
            )
        i += 1
        previous_height = height

    return ax


def scree_plot(pca, out_folder=None, **kwargs):
    """
    Scree plot of explained variance from each principal components of a PCA object.
    Parameters
    ----------
    pca : sklearn.decomposition.PCA
        a pre-fitted PCA model from sklearn
    out_folder : str, default=None
        Output folder where to save the plot. File name will be scree-plot.png
    title : str, OPTIONAL
        Title of the scree plot
    Returns
    -------
    fig : matplotlib.pyplot.figure
        Figure Object of the plot
    ax : matplotlib.pyplot.axes
        Axes Object of the plot
    """

    # Figure set up
    n = pca.n_components_
    evr = pca.explained_variance_ratio_
    fig_width = len(evr)
    fig, ax = plt.subplots(figsize=(12, 9))
    width = 0.02 * fig_width

    # Line plot
    x_plot = range(1, fig_width + 1)
    y_plot = 100 * np.cumsum(evr)
    plt.plot(x_plot, y_plot, c='red', label='Cumulative Explained Variance')

    # Bar plot
    x_bar = x_plot
    y_bar = 100 * evr
    plt.bar(x_bar, y_bar, width=width)

    # Error bar
    yerr = np.cumsum(100 * evr) - 100 * evr
    err = ax.errorbar(x_bar, y_bar, yerr=yerr, lolims=True, capsize=0,
                      ls='None', color='k', elinewidth=0.1)

    # Title and labels
    ax.set_title('PCA Scree Plot', fontsize=14, fontweight='bold', pad=10)
    ax.set_xlabel('Number of components', fontsize=12, fontweight='bold')
    ax.set_ylabel('Explained variance (%)', fontsize=12, fontweight='bold')
    ax = _autolabel( np.cumsum(100 * evr), n, ax)
    #_autolabel(np.cumsum(pca.explained_variance_ratio_ * 100), fig_width, ax)


    plt.xticks(range(1, n + 1), fontsize=9)
    plt.yticks(range(0, 110, 10), fontsize=10)
    plt.legend(loc='center right', framealpha=1, fontsize=10)
    if(out_folder != None):
        if (os.path.exists(out_folder) != True):
            os.makedirs(out_folder)
            
        if 'title' in kwargs:
            # We cannot fix the title as there will be as many scree plots as the number of Thematic PCA domains.
            # I (Ozer) am making the title optional using kwargs so that we can create scree plots as the number of domains in the Thematic PCA template 
            # and save them by giving them different names.
            plt.savefig(os.path.join(out_folder, kwargs['title']))
        else:
            plt.savefig(os.path.join(out_folder, 'scree_plot.png'))

    return fig, ax

def correlation_matrix_plot(df, output_path, corr_threshold=None, method='spearman', **kwargs):
    """
    Plot a Spearman correlation matrix to assess linear correlation between each columns in a DataFrame
    Parameters
    ----------
    df : pandas.DataFrame
        input data
    output_path : str
        Path where to save a excel containing the correlation coefficient
    corr_threshold : float
        Threshold to use. It keeps and plots only variables with correlation above given threshold.
    method :{'pearson', 'kendall', 'spearman'} or callable
        Method of correlation, as per pandas.DataFrame.corr
    **kwargs : optional seaborn.heatmap keyword arguments    
     
    Examples
    ----------
    correlation_matrix_plot(df_ica, output_folder, method='spearman')
    correlation_matrix_plot(df_ica, output_folder, corr_treshold=0.60, method='spearman') 
    correlation_matrix_plot(df_ica, output_folder, corr_treshold=0.60, method='spearman', annot=True) 
    
    Returns
    -------
    None
    """
    method_title = str(method).capitalize()
    cmap = mpl.cm.Reds    

    fig, ax = plt.subplots(figsize=(100, 80))
    
    if corr_threshold == None:
        corr = df.corr(method=method).abs()
        mask = np.triu(np.ones_like(corr))     
    else:
        corr = df.corr(method=method).abs()
        corr = corr[(corr >= corr_threshold) & (corr !=1.000)]
        mask = np.triu(np.ones_like(corr, dtype=bool))
    
    ax = sns.heatmap(corr,
                ax = ax,
                cmap=cmap, vmin=0.0, vmax=1.0, linewidths=10, linecolor='w',
                annot_kws={'size': 50, 'weight': 'bold'}, fmt='.2f', 
                square=True, mask=mask,
                cbar=True,
                cbar_kws={
                    'pad': .01, 
                    'ticks': [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0],
                    'shrink': .82,
                    'extend': 'both'
                         },  
                **kwargs)
    
    cax = plt.gcf().axes[-1]
    cax.tick_params(labelsize=120)
    
    # Add Tick Labels
    labels = [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
    cax.set_yticklabels(labels, fontsize=80, weight='bold')
    cax.set_ylabel(f'\n\n{method_title} Correlation', fontsize=80, weight='bold')
    
    # Setup Ticks and Labels
    plt.title(f"Features Heat Map with {method_title} Correlation\n", pad=30, fontsize=120, fontweight='bold')
    plt.xlabel('\n\nFeatures', labelpad=30, fontsize=120, fontweight='bold')  # x-axis label with fontsize 15
    plt.ylabel('Features\n\n', labelpad=30, fontsize=120, fontweight='bold')  # y-axis label with fontsize 15
    plt.xticks(fontsize=60, fontweight='bold', rotation=90)
    plt.yticks(fontsize=60, fontweight='bold', rotation=0)
    plt.tight_layout()
    fig.savefig(os.path.join(output_path, f'{method_title}_Correlations.png'), dpi=100)    
    plt.show()
    
    if not os.path.exists(output_path):
        os.makedirs(output_path)

    pd.DataFrame(corr).to_excel(os.path.join(output_path, "correlation_matrix.xlsx"))

    return fig, ax

def pca_dataframe(dataframe, coordinates_index, explained_variance=None, scaler='standard_scaler', random_state=42, output_folder=None, thematic_pca=True, columns=None, save_df_pca=True, no_data_value=None):
    """
    Apply Thematic PCA on specific subset of the cube and return the reduced pc transformed values in a DataFrame.
    It is suggested that dataframe should be either standard-scaled or robust-scaled before.
    Parameters
    ----------
    dataframe : pandas.DataFrame or dask.DataFrame
        DataFrame containing the numerical data to be sent to a PCA.fit() method.
    coordinates_index : array
        Row index for of the X and Y coordinates. Idea is that you can use pandas.join() on indexes on the dataframe returned. If left to None, indexes will be 0,1,2,3...,n
    explained_variance : float or int, default=None
        n_components argument for PCA.
        Any float number between 0 and 1 if it is used for explained variance.
        However, it is integer for any values larger than 1 if it is used for keeping the number of PC columns in the function
    scaler : str, default='standard_scaler'
        Name of the scaling strategy to use. By default, standard_scaler is applied. Other available method is robust_scaler.
    random_state : int, default=42
        Random seed for PCA
    output_folder : string, default=None
        Output folder where to save the DataFrame. Its name will be PC_Transformed_DataFrame.csv. If left to None, it does not get saved.
    thematic_pca : boolean, OPTIONAL, default=True
        Performs thematic PCA if it is set thematic_pca = True (by default)
        Performs normal PCA for the given pandas dataframe
    columns : dict, columns dictionary of datawrangle output or columns dictionary given to the HyperCube class.
        OR path to the hypercube columns in json format.
        Mandatory to perform thematic PCA. If the columns are not assigned by the user, the function automatically performs normal PCA even if thematic_pca is set to True.
        datawrangle.cube_vrt(..., save_columns=True) function saves a JSON format of the columns names used while building the cube.
        The columns object will also be allowed to be a JSON file, which is already given by datawrangle.cube_vrt(..., save_columns=True) function.
    Returns
    -------
    df_pca : pandas.DataFrame
        DataFrame of the PCs transformed values. Columns will be named PC1, PC2, PC3, etc...
    pca or thematic_pca_objects: sklearn.decomposition.PCA or a dictionary of thematic_pca_objects, respectively
        Fitted PCA object for single PCA or Fitted PCA objects for Thematic PCA, respectively
    """

    if no_data_value != None:
        for column in dataframe.columns: dataframe[column] = dataframe[column].replace(no_data_value, np.nan)
    
    #dataframe = dataframe[columns['coordinates'] + columns['numerical']].copy()
    dataframe = dataframe.dropna() # In case dataframe has some Nan values let's drop them!
    
    # If the given dataframe contains coordinates, let's drop them.
    columns_to_drop = ['x', 'y', 'z', 'X', 'Y', 'Z'] # Possible coordinates
    df_coord = dataframe[dataframe.columns.intersection(columns_to_drop).tolist()] # coordinates dataframe
    dataframe = dataframe[dataframe.columns.difference(columns_to_drop).tolist()] # dataframe without coordinates
                
    # Let's find the thematic PCA domains!
    dict_domain_df = {} # The dictionary to host the dataframes that are slided for each thematic pca domain
    dict_domain_pca = {} # The dictionary of pca results for each thematic pca

    # Possible target key names that the user can set in the columns dictionary
    list_targets_names = ['Target', 'target', 'Targets', 'targets']
    # Possible coordinate key names that the user can set in the columns dictionary
    list_coordinate_names = ['coordinates','coordinates'.capitalize()]
    # Possible categorical key names that the user can set in the columns dictionary
    list_categorical_names = ['Categorical', 'categorical', 'Categories', 'categories', 'Category', 'category']
    # Possible numerical key names that the user can set in the columns dictionary
    list_numerical_names = ['Numerical', 'numerical', 'Num', 'num', 'Numeric', 'numeric']
    # Possible features key names that the user can set in the columns dictionary to see all features together
    list_feature_names = ['Features', 'features']

    # Domains to which Thematic PCA will not be applied under any circumstances!
    excluded_domains = list_coordinate_names + list_categorical_names + list_numerical_names + list_feature_names + list_targets_names

    if columns != None:
        if type(columns) == dict:
            thematic_pca_domains = [column for column in columns.keys() if column not in excluded_domains] # Domains to which PCA will be applied
        elif type(columns) == str:
            file, extension, directory = utilities.Path_Info(columns)
            if (extension == '.json'):
                # If columns is the path to the json file
                with open(columns, 'r') as json_file: json_dict = json.load(json_file)   
                thematic_pca_domains = [column for column in json_dict.keys() if column not in excluded_domains] # Domains to which PCA will be applied
            else:
                thematic_pca_domains = []
    else:
        thematic_pca_domains = []
        
    PC_columns_dict = {} # Renamed PC columns for Thematic PCA
    thematic_pca_objects = {} # Equivalent of pca class for normal PCA but this one includes pca classes for each domain        
            
    scalers = ['robust_scaler', 'standard_scaler']
    if(scaler in scalers):
        if(scaler=='robust_scaler'):
            scaler = RobustScaler(with_centering=True,
                                  with_scaling=True,
                                  quantile_range=(25.0, 75.0),
                                  copy=True)
        elif(scaler=='standard_scaler'):
            scaler = StandardScaler()

        if (thematic_pca == True and columns != None): 
            if type(columns)==dict:
                # If the user wants to apply thematic PCA and hypercube columns dictionary is given to the function!                     
                for pca_domain in thematic_pca_domains:
                    dict_domain_df[pca_domain] = dataframe[columns[pca_domain]] # Let's set the dataframes to the dictionary
                    dict_domain_df[pca_domain] = scaler.fit_transform(dict_domain_df[pca_domain]) # Let's apply scaling for each domain     
                    dict_domain_df[pca_domain] = pd.DataFrame(dict_domain_df[pca_domain], columns=dataframe[columns[pca_domain]].columns) # Convert np.array to dataframe   

            elif type(columns)==str:
                file, extension, directory = utilities.Path_Info(columns)
                if (extension == '.json'):
                    # If the user wants to apply thematic PCA and columns is the path to the json file!
                    with open(columns, 'r') as json_file: json_dict = json.load(json_file)
                    for pca_domain in thematic_pca_domains:
                        dict_domain_df[pca_domain] = dataframe[json_dict[pca_domain]] # Let's set the dataframes to the dictionary
                        dict_domain_df[pca_domain] = scaler.fit_transform(dict_domain_df[pca_domain]) # Let's apply scaling for each domain     
                        dict_domain_df[pca_domain] = pd.DataFrame(dict_domain_df[pca_domain], columns=dataframe[json_dict[pca_domain]].columns) # Convert np.array to dataframe   
                         
        else:
            # If the user only wants to apply normal PCA (and not thematic PCA), this else block will take care of it.
            # Set thematic_pca = False
            
            # If the given dataframe contains coordinates, let's drop them and then fit transform
            dataframe = scaler.fit_transform(dataframe.drop(columns=dataframe.columns.intersection(columns_to_drop))) # fit transform for dataframe without coordinates
            
    else:
        if(scaler == None or scaler == 'none'):
            print("No scaling will be performed.")
            if type(columns)==dict:
                for pca_domain in thematic_pca_domains:
                    dict_domain_df[pca_domain] = dataframe[columns[pca_domain]] # Let's set the dataframes to the dictionary
            elif type(columns) == str:
                file, extension, directory = utilities.Path_Info(columns)
                if (extension == '.json'):
                    # If columns is the path to the json file
                    with open(columns, 'r') as json_file: json_dict = json.load(json_file)
                    for pca_domain in thematic_pca_domains:
                        dict_domain_df[pca_domain] = dataframe[json_dict [pca_domain]] # Let's set the dataframes to the dictionary
            else:
                pass
        else:
            raise (ValueError("The scaling method "))

    if (thematic_pca == True and columns != None):         
        # If thematic_pca = True and the columns is provided!
        # Thematic PCA will be performed in this if block!

        for key in dict_domain_df.keys():
            print('Performing Thematic PCA for the domain of ', key)
            pca = PCA(random_state=random_state, n_components = explained_variance) # We want the explained variance to be 99%
            dict_domain_pca[key] = pca.fit_transform(dict_domain_df[key]) # PCA fit transform for specific domains
            original_dim = dict_domain_df[key].shape[1] # Dimension of the original data
            reduced_dim = dict_domain_pca[key].shape[1] # Dimension of the reduced data
            print('The dimension is reduced from ', original_dim, ' to ', reduced_dim, f' with the expected variance of {explained_variance}')            
            PC_columns_dict.update({key : [f'PC{x + 1}_'+str(key) for x in range(reduced_dim)]}) # PC column names   
            dict_domain_pca[key] = pd.DataFrame(dict_domain_pca[key], columns=PC_columns_dict[key]) # Convert np.array to dataframe 
            thematic_pca_objects.update({key : pca}) # Equivalent of pca objects for normal PCA but this one includes pca objects for each domain     
        
        df_pca = pd.concat([dict_domain_pca[key] for key in dict_domain_pca], axis=1) # Concatenate thematic PCA results in a single dataframe
        # Let's concat df_coord and df_pca to get a PC Transformed DataFrame Including Coordinates
        df_pca.reset_index(inplace=True, drop=True)
        df_coord.reset_index(inplace=True, drop=True)
        df_pca = pd.concat([df_coord, df_pca], axis=1) # PC Transformed DataFrame Including Coordinates 
        
    else:
        # Possible scenarios:
        # If thematic_pca = False 
        # OR thematic_pca = True but the user forgets to set a columns dictionary 
        # Normal PCA will be performed in this else block!
                
        pca = PCA(random_state=random_state, n_components = explained_variance)        
        pc_array = pca.fit_transform(dataframe)
        df_pca = pd.DataFrame(pc_array, columns=_get_pcs_labels(pc_array.shape[1]), index=coordinates_index)
        original_dim = dataframe.shape[1] # Dimension of the original data
        reduced_dim = df_pca.shape[1]     # Dimension of the reduced data
        print('The dimension is reduced from ', original_dim, ' to ', reduced_dim, f' with the expected variance of {explained_variance}')
        # Let's concat df_coord and df_pca to get a PC Transformed DataFrame Including Coordinates
        df_pca.reset_index(inplace=True, drop=True)
        df_coord.reset_index(inplace=True, drop=True)
        df_pca = pd.concat([df_coord, df_pca], axis=1) # PC Transformed DataFrame Including Coordinates 
        
    if(output_folder != None and save_df_pca == True):
        print("Saving the PC transformed data as a csv... it may take a while")
        df_pca.to_csv(os.path.join(output_folder, 'PCA_Transformed_DataFrame.csv'), index=True)

    if (thematic_pca == True and columns != None):
        return df_pca, thematic_pca_objects
    else:
        return df_pca, pca


def ica_dataframe(numeric_data,  coordinates_index, output_directory=None,random_state=42):
    ica = FastICA(random_state=random_state)
    ica_array = ica.fit_transform(numeric_data)
    n_variables = numeric_data.shape[1]
    df_ica = pd.DataFrame(ica_array, columns=[f'IC{x + 1}' for x in range(n_variables)], index=coordinates_index)

    if (output_directory != None):
        if not os.path.exists(output_directory):
            os.makedirs(output_directory)

        df_ica.to_csv(os.path.join(output_directory, 'ICA_Transformed_DataFrame.csv'), index=True)

    return df_ica, ica


def plot_pc_loadings(pca_transformer, feature_list, output_directory, dpi=300, **kwargs):
    """
    Plot loadings for each individual principal components of a fitted PCA Transformer. X will be the loading, Y will be the features (original variables).
    Function taken from Shawn's MapClass repository.
    Parameters
    ----------
    pca_transformer : sklearn.decomposition.PCA
        Fitted PCA Transformer object
    feature_list : list
        List of the original variables name. They must have the same order that the DataFrame's columns that was sent to fit the PCA transformer.
    output_directory : string
        Output directory where to save the images.
    dpi : int
        Resolution (dots per inch). Leave it to the default value. default=300
    thematic_domain : str, OPTIONAL
        Thematic domain name
        This is required for src/GeoDS/scripts/preprocessing/thematic_pca.py to set proper file names
    Returns
    -------
    fig : Matplotlib.Figure
        Figure from matplotlib
    ax : Matplotlib.Axe
        Axe object from matplotlib
    """
    
    PCcolnames = _get_pcs_labels(pca_transformer.n_components_)

    loadings_orig = pd.DataFrame(
        data=pca_transformer.components_.T, index=feature_list, columns=PCcolnames)      
        
    loadings_sorted = pd.DataFrame()
    for i in loadings_orig.columns[:]:
        if 'thematic_domain' in kwargs:
            print(str(kwargs['thematic_domain'])+' '+str(i))
        else:
            print(i)
        # get relevant columns and sort by PC
        _ = loadings_orig[i].reset_index().sort_values(by=[i], ascending=False)
        # Concatenate a table for export
        loadings_sorted = pd.concat(
            [loadings_sorted, _.reset_index(drop=True)], axis=1
        )
        # Plot the figure
        fig = plt.figure()
        ax = plt.axes()
        x = _["index"]
        y = _[i]
        ax.plot(-y, x, "ko-", linewidth=1, markersize=4)
        plt.xticks((-0.5, 0, 0.5), rotation=0)
        # plt.legend()
        ax.text(
            0.05,
            0.95,
            i,
            transform=ax.transAxes,
            fontsize=9,
            verticalalignment="top",
        )
        plt.tight_layout()
        
        if 'thematic_domain' in kwargs:
            # This block is required for src/GeoDS/scripts/preprocessing/thematic_pca.py
            # I am making thematic_domain optional using kwargs so that I can create pc_loadings plots as the number of domains in the Thematic PCA template 
            # and save them by setting them proper names as below!
            plt.savefig(f"{output_directory}{kwargs['thematic_domain']}_{i}_loading.png", format="png", dpi=dpi)
        else:
            plt.savefig(f"{output_directory}{i}_loading.png", format="png", dpi=dpi)
        
        plt.show()

    return fig, ax


def plot_loadings_heatmap(pca_transformer, original_labels, output_directory=None, **kwargs):
    """
    Plot a heatmap to visualize loadings vs labels.
    Parameters
    ----------
    pca_transformer : sklearn.decomposition.PCA
        Fitted PCA Transformer object
    original_labels : list
        Original variables names
    output_directory : str
        Directory where to save the plot. Its name will be 'eigenvectors-loadings-heatmap.png'. If left to None, will not be saved.
    title : str, OPTIONAL
        Name of the heatmap plot
    Returns
    -------
    fig : Matplotlib.Figure
        Figure from matplotlib
    ax : Matplotlib.Axe
        Axe object from matplotlib
    """
    n = pca_transformer.n_components_
    compo = pca_transformer.components_
    pc_lstr = _get_pcs_labels(n)
    df = pd.DataFrame(data=compo, index=pc_lstr, columns=original_labels)

    #fig, ax = plt.subplots(figsize=(n / 4, n / 4))
    size_x = 2/3 * n
    size_y = 2/3 * len(original_labels)

    fig, ax = plt.subplots(figsize=(size_x, size_y))
    norm = TwoSlopeNorm(vmin=-1, vcenter=0, vmax=1)
    im = ax.imshow(compo, cmap='seismic_r', norm=norm)
    divider = make_axes_locatable(ax)
    cax = divider.append_axes('right', size=0.25, pad=0.075)
    plt.colorbar(im, cax=cax)

    # Title and labels
    ax.set_title(
        'Principal components composition',
        fontsize=14,
        fontweight='bold',
        pad=10
    )
    ax.set_xticks(np.arange(n))
    ax.set_xticklabels(original_labels, rotation=45, ha='right', fontsize=8)
    ax.set_yticks(np.arange(n))
    ax.set_yticklabels(pc_lstr)

    plt.tight_layout()
    if (output_directory != None):
        if not os.path.exists(output_directory):
            os.makedirs(output_directory)

        if 'title' in kwargs:
            # We cannot fix the title as there will be as many heatmap plots as the number of Thematic PCA domains.
            # I (Ozer) am making the title optional using kwargs so that we can create heatmap plots as the number of domains in the Thematic PCA template 
            # and save them by giving them different names.
            plt.savefig(os.path.join(output_directory, kwargs['title']))
        else:
            plt.savefig(os.path.join(output_directory, 'eigenvectors-loadings-heatmap.png'))
            
    # Figure save
    return fig, ax

def _get_loadings_dataframe(pca_transformer, original_labels):
    n = pca_transformer.n_components_
    compo = pca_transformer.components_
    pc_lstr = _get_pcs_labels(n)
    df = pd.DataFrame(data=compo, index=pc_lstr, columns=original_labels)
    return df

def loadings_dataframe(pca_transformer, original_labels, output_directory=None):
    """
    Get loadings of eigenvectors into a DataFrame
    Parameters
    ----------
    pca_transformer : sklearn.decomposition.PCA
        Fitted PCA Transformer object
    original_labels : list
        Name of the original variables. They must have the same order that the DataFrame's columns that was sent to fit the PCA transformer.
    output_directory : string, default=None
        Where to save the dataframe. Its name will be 'eigenvectors-loadings.csv'. If left to None, will not be saved.
    Returns
    -------
    df : pandas.DataFrame
        dataframe that contains the loadings for each principal components.
    """
    df = _get_loadings_dataframe(pca_transformer, original_labels)

    # Set up
    if(output_directory != None):
        if not os.path.exists(output_directory):
            os.makedirs(output_directory)

        df.to_csv(os.path.join(output_directory, 'eigenvectors-loadings.csv'))

    return df


def _columns_adjuster(worksheet):
    """
    Private function that helps to adjust the width of the cells while building the evaluation template.
    """
    # credit to this dude https://stackoverflow.com/questions/39529662/python-automatically-adjust-width-of-an-excel-files-columns
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            cell.alignment = Alignment(horizontal='center')
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = (max_length + 2) * 1.1
        # original was +2 * 1.2 but that was a bit too wide.
        worksheet.column_dimensions[column].width = adjusted_width

    return worksheet

    # https://towardsdatascience.com/reading-and-visualizing-geotiff-images-with-python-8dcca7a74510

def pc_evaluation_template(pca, original_features_names, output_directory, jpgs_folder=None, **kwargs):
    """
    Generate MacLean's PC Evaluation Template with adequate symbology and images. Each of the principal component is assessed visually and the geologist / geodatascientist decide if the information captured by the PC is relevant, a.k.a. if we need to keep or drop the principal component as a feature for model building.
    Parameters
    ----------
    pca : sklearn.decomposition.PCA
        pre-fitted pca object
    original_features_names : list
        original names of your variables (features)
    output_directory : str
        path to the output directory
    jpgs_folder : str, default=None
        path to the jpgs folder of the principal component. Each of the PC should be named PC1.jpg, PC2.jpg, PC3.jpg. If you used all the GeoDS tools. Default = None. No image will be added to the excel sheets.
    thematic_domain : str, OPTIONAL
        Correction to the title of the workbook in case of thematic PCA     
    Returns
    -------
    None
    """
    # if user provided jpgs_folder, we will add plots to each individual sheet of the principal components
    #Lets start by making a dictionary that will look like {'PC1': 'my_jpg_folder/PC1.tif'}
    if (jpgs_folder != None):
        pcs_images_dict = {}
        jpg_list = glob.glob(jpgs_folder + '*.jpg')

        for j in jpg_list:
            filename, extension, directory = utilities.Path_Info(j)
            # raster_to_jpg(t, filename + '.jpg')
            pcs_images_dict[filename] = j

    #Pre-define the colorscale to format the loadings
    rule = ColorScaleRule(start_type='percentile', start_value=5, start_color='FFF8696B',
                          mid_type='percentile', mid_value=50, mid_color='FFFFEB84',
                          end_type='percentile', end_value=95, end_color='FF63BE7B')

    #columns for the summary A, B, C,D,E,F,G,H,I,J
    columns = ['Principal Component', 'Geological', 'Appears Random', 'Survey Noise', 'Topography', 'Water', 'Power',
               'Other', 'Comments', 'Keep/Drop', 'Evaluated By']

    #pca must be pre-fitted
    n = pca.n_components_
    
    #name of the PC columns ['PC1', 'PC2', etc.]
    pc_lstr = _get_pcs_labels(n)

    #Create a semi-empty DataFrame with right columns for the Summary sheet.
    observations_df = pd.DataFrame(columns=columns)
    observations_df['Principal Component'] = pc_lstr

    #Create the DataFrame from the loadings of each PC vs original features for the Loadings sheet
    loadings_df = loadings_dataframe(pca, original_features_names)
    loadings_df_out = loadings_df.reset_index()
    loadings_df_out.rename({'index': 'Principal Component'}, axis=1, inplace=True)

    #Create a Workbook object
    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'

    #Write the Summary sheet. Add picklists in cells.
    for r in dataframe_to_rows(observations_df, index=False, header=True):
        ws.append(r)
    ws = _columns_adjuster(ws)
    data_val = DataValidation(type="list", formula1='"Yes,No,Maybe,Not Sure"', allow_blank=True)
    data_val2 = DataValidation(type="list", formula1='"Keep,Drop,TBD"', allow_blank=True)
    ws.add_data_validation(data_val)
    ws.add_data_validation(data_val2)
    data_val.add("B2:G" + str(n + 1))
    data_val2.add("J2:J" + str(n + 1))

    #Write Loadings sheet. Add the colorscale formatting.
    ws2 = wb.create_sheet(title='Loadings')
    for r in dataframe_to_rows(loadings_df_out, index=False, header=True):
        ws2.append(r)
    ws2 = _columns_adjuster(ws2)
    # https://stackoverflow.com/questions/12902621/getting-the-row-and-column-numbers-from-coordinate-value-in-openpyxl
    letter = get_column_letter(n + 1)
    last_cell = letter + str(n + 1)
    ws2.conditional_formatting.add('A1:' + last_cell, rule)

    #Create a sub DataFrame for each PC by first getting the loadings from loadings_df.
    #Sort the loadings the decreasing way.
    #Add an empty comment column.
    for z in pc_lstr:
        subdf = loadings_df.loc[z].to_frame(name=z)
        subdf = subdf.sort_values(by=z, ascending=False)
        subdf.reset_index(inplace=True)
        subdf.rename({'index': 'Original Feature'}, axis=1, inplace=True)
        subdf['Comment'] = pd.Series(dtype='int')

        #write the sheet for the specific PC
        ws_temp = wb.create_sheet(title=z)
        for r in dataframe_to_rows(subdf, index=False, header=True):
            ws_temp.append(r)
        ws_temp = _columns_adjuster(ws_temp)

        # The following if block creates an issue for src/GeoDS/scripts/preprocessing/thematic_pca.py
        # I have created the PCA_Evaluation_Template without these lines. Everything looks fine! 
        # We don't need these lines just to make the template look a little better.
        #if (jpgs_folder != None):
        #    img = Image(pcs_images_dict[z])
        #    ws_temp.add_image(img, 'F1')

        #add the colorscale to the loading
        ws_temp.conditional_formatting.add('B1:B' + str(n + 1), rule)

    # create formulas so that the comment column is filled with the comments from cell C2 (comment cell) in each of the PC sheets
    for i in np.arange(1, n + 1, 1):
        condition = "'PC" + str(i) + "'!C2"
        ws["I" + str(i + 1)] = "=IF(" + condition + " = \"\", \"\"," + condition + ")"

    if 'thematic_domain' in kwargs:
        # This block is required for src/GeoDS/scripts/preprocessing/thematic_pca.py
        # I am making thematic_domain optional using kwargs so that I can create pc evaluation templates as the number of domains in the Thematic PCA template 
        # and save them by setting them proper names as below!     
        wb.save(os.path.join(output_directory, 'PCA_Evaluation_'+kwargs['thematic_domain']+'.xlsx' ))
    else:
        wb.save(os.path.join(output_directory, 'PCA_Evaluation_Template.xlsx'))
    
    return None


def ic_evaluation_template(ica, original_features_names, output_directory, jpgs_folder=None):
    if (jpgs_folder != None):
        pcs_images_dict = {}
        jpg_list = glob.glob(jpgs_folder + '*.jpg')

        for j in jpg_list:
            filename, extension, directory = utilities.Path_Info(j)
            # raster_to_jpg(t, filename + '.jpg')
            pcs_images_dict[filename] = j



    #columns for the summary A, B, C,D,E,F,G,H,I,J
    columns = ['Independent Component', 'Geological', 'Appears Random', 'Survey Noise', 'Topography', 'Water', 'Power',
               'Other', 'Comments', 'Keep/Drop', 'Evaluated By']

    #pca must be pre-fitted
    n = ica.components_.shape[0]
    #name of the PC columns ['PC1', 'PC2', etc.]
    pc_lstr = _get_ics_labels(n)

    #Create a semi-empty DataFrame with right columns for the Summary sheet.
    observations_df = pd.DataFrame(columns=columns)
    observations_df['Independent Component'] = pc_lstr

    #Create the DataFrame from the loadings of each PC vs original features for the Loadings sheet
    #loadings_df = loadings_dataframe(pca, original_features_names)
    #loadings_df_out = loadings_df.reset_index()
    #loadings_df_out.rename({'index': 'Principal Component'}, axis=1, inplace=True)

    #Create a Workbook object
    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'

    #Write the Summary sheet. Add picklists in cells.
    for r in dataframe_to_rows(observations_df, index=False, header=True):
        ws.append(r)
    ws = _columns_adjuster(ws)
    data_val = DataValidation(type="list", formula1='"Yes,No,Maybe,Not Sure"', allow_blank=True)
    data_val2 = DataValidation(type="list", formula1='"Keep,Drop,TBD"', allow_blank=True)
    ws.add_data_validation(data_val)
    ws.add_data_validation(data_val2)
    data_val.add("B2:G" + str(n + 1))
    data_val2.add("J2:J" + str(n + 1))

    #Write Loadings sheet. Add the colorscale formatting.
    #ws2 = wb.create_sheet(title='Loadings')
    #for r in dataframe_to_rows(loadings_df_out, index=False, header=True):
    #    ws2.append(r)
    #ws2 = _columns_adjuster(ws2)
    ## https://stackoverflow.com/questions/12902621/getting-the-row-and-column-numbers-from-coordinate-value-in-openpyxl
    #letter = get_column_letter(n + 1)
    #last_cell = letter + str(n + 1)
    #ws2.conditional_formatting.add('A1:' + last_cell, rule)

    #Create a sub DataFrame for each PC by first getting the loadings from loadings_df.
    #Sort the loadings the decreasing way.
    #Add an empty comment column.
    for z in pc_lstr:
        subdf = pd.DataFrame(columns=['Comment'])
        #subdf = loadings_df.loc[z].to_frame(name=z)
        #subdf = subdf.sort_values(by=z, ascending=False)
        #subdf.reset_index(inplace=True)
        #subdf.rename({'index': 'Original Feature'}, axis=1, inplace=True)
        #subdf['Comment'] = pd.Series(dtype='int')

        #write the sheet for the specific PC
        ws_temp = wb.create_sheet(title=z)
        for r in dataframe_to_rows(subdf, index=False, header=True):
            ws_temp.append(r)
        ws_temp = _columns_adjuster(ws_temp)

        #get the corresponding
        if (jpgs_folder != None):
            img = Image(pcs_images_dict[z])
            ws_temp.add_image(img, 'F1')

        #add the colorscale to the loading
        #ws_temp.conditional_formatting.add('B1:B' + str(n + 1), rule)

    # create formulas so that the comment column is filled with the comments from cell C2 (comment cell) in each of the PC sheets
    for i in np.arange(1, n + 1, 1):
        condition = "'IC" + str(i) + "'!A2"
        ws["I" + str(i + 1)] = "=IF(" + condition + " = \"\", \"\"," + condition + ")"

    wb.save(os.path.join(output_directory, 'ICA_Evaluation_Template.xlsx'))
    return None


def normalize_by_row(matrix):
    """ Make every row-vector unit.
    """
    m, n = np.shape(matrix)
    L2 = norm(matrix, axis=1).reshape((m, 1))
    return matrix / L2

def components_assess(ica, out_folder, original_labels):
    """ Eigenvectors heatmap and csv export.
    Note that every row-vector is independently renormalized to unit vector for
    visualization purposes, while the csv is kept intact.
    """
    # Set up
    compo = ica.components_
    n, nfeat = np.shape(compo)
    ic_lstr = [f'IC{x + 1}' for x in range(n)]
    compo_norm = normalize_by_row(compo)

    # Heatmap plot
    fig, ax = plt.subplots(figsize=(18.5, 10.5))
    fig.set_facecolor('white')
    norm = TwoSlopeNorm(vmin=-1, vcenter=0, vmax=1)
    im = ax.imshow(compo_norm, cmap='seismic_r', norm=norm)
    divider = make_axes_locatable(ax)
    cax = divider.append_axes('right', size=0.25, pad=0.075)
    plt.colorbar(im, cax=cax)

    # Title and labels
    ax.set_title(
        'Independant components composition\n(independently renormalized to unit)',
        fontsize=14,
        fontweight='bold',
        pad=10
    )
    ax.set_xticks(np.arange(n))
    ax.set_xticklabels(original_labels, rotation=45, ha='right', fontsize=8)
    ax.set_yticks(np.arange(n))
    ax.set_yticklabels(ic_lstr)

    # Figure save
    plt.tight_layout()
    plt.savefig(os.path.join(out_folder,'components-heatmap.png'))

    # Write csv
    df = pd.DataFrame(data=compo, index=ic_lstr, columns=original_labels)
    df.to_csv(os.path.join(out_folder, 'components-compo.csv'))


